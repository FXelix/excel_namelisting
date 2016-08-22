
# open a .txt, search for names and print them to an excel document.

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.cell import get_column_letter
import re
import codecs

# regex: forename and surname
name_regex = r"[A-Za-zäüöÖÜÄ]+\s+" \
             r"[A-Za-zäüöÖÜÄ]+\s*"

name_compiled = re.compile(name_regex, re.VERBOSE)

#open the excel document
wb = openpyxl.load_workbook("namelist.xlsx")  # here goes your excel-file
sheet = wb.get_sheet_by_name("Namensliste")   # here goes the sheet-name

# styling all borders for readability
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# open file with names
f = codecs.open("names.txt", encoding="utf-8")
name = f.readlines()
name = "".join(name)
print(" Raw names ".center(40, "="), "\n")

#find names that match the pattern
raw_names = name_compiled.findall(name)
raw_names = sorted(raw_names, key=lambda s: s.lower())  # sort them alphabetically
print(raw_names)
print(" All names ".center(40, "="))

# info about names and lines
count_names = len(raw_names)
line_count = sum(1 for line in open('names.txt'))
print("{} names found in {} lines!".format(count_names, line_count))
if count_names == line_count:
    print("INFO: Perfect accordance!")
else:
    print("INFO: Less names than lines in the document were found..")
print("-" * 40)

r = 1   # starting row 1
c = 1   # starting column A

# iterate over all names found and print them into the excel document
for n in raw_names:
    cell_ort = sheet.cell(row=r, column=c).coordinate
    print("Added:", n, end="")
    sheet[cell_ort] = n

    sheet[cell_ort].border = thin_border  # adjust all borders as definded above
    col_letter = get_column_letter(c)     # get the column letter e.g "B"
    sheet.column_dimensions[col_letter].width = 22  # adjusting width of column
    r += 1
    if r == 51:  # if you reached the end of the printing-page
        c += 1   # next column
        r = 1    # next row

#  quit and save
wb.save("namelist.xlsx")
f.close()
